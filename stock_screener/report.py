"""
Formats and exports screener results.

Outputs:
  - Rich terminal table (ranked by score)
  - results/scores.csv       — full data for every stock
  - results/scores.xlsx      — Excel with conditional formatting
  - results/top20_chart.png  — bar chart of top 20 stocks
  - results/heatmap.png      — score category heatmap
"""

import json
import os
import sys
from datetime import datetime

import numpy as np
import pandas as pd

from rich.console import Console
from rich.table import Table
from rich import box
from rich.panel import Panel

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from stock_screener.scorer import ScoreBreakdown

console = Console()

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    MATPLOTLIB = True
except ImportError:
    MATPLOTLIB = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL = True
except ImportError:
    EXCEL = False


# ─── Terminal display ─────────────────────────────────────────────────────────

def _signal_style(signal: str) -> str:
    return {
        "STRONG BUY": "bold green",
        "BUY":        "green",
        "WATCH":      "yellow",
        "NEUTRAL":    "dim",
        "AVOID":      "red",
    }.get(signal, "white")


def _score_color(score: float, max_score: float) -> str:
    pct = score / max_score
    if pct >= 0.80: return "bold green"
    if pct >= 0.60: return "green"
    if pct >= 0.40: return "yellow"
    return "red"


def _fmt(v, fmt=".1f", suffix="", na="—"):
    """Format a float, returning na string if NaN."""
    try:
        import math
        if v is None or math.isnan(float(v)):
            return na
        return f"{float(v):{fmt}}{suffix}"
    except (TypeError, ValueError):
        return na


def print_results(scores: list[ScoreBreakdown], top_n: int = 30):
    valid = [s for s in scores if not s.error]
    valid.sort(key=lambda s: s.total, reverse=True)
    show  = valid[:top_n]

    timestamp = datetime.now().strftime("%d %b %Y  %H:%M")
    console.print()
    console.print(Panel(
        f"  [bold white]NSE / BSE STOCK SCREENER — FUNDAMENTAL ANALYSIS[/bold white]   "
        f"[dim]{len(valid)} stocks scored  |  {timestamp}[/dim]",
        border_style="blue", box=box.DOUBLE_EDGE,
    ))
    console.print()

    t = Table(
        title=f"Top {len(show)} Stocks by Fundamental Score",
        box=box.SIMPLE_HEAVY, border_style="blue",
        title_style="bold white", show_lines=False,
    )

    t.add_column("Rank",      justify="right",  min_width=4)
    t.add_column("Symbol",    style="bold cyan", min_width=12)
    t.add_column("Name",                         min_width=22)
    t.add_column("Exch/Cap",  justify="center",  min_width=9)
    t.add_column("Price ₹",   justify="right",   min_width=9)
    t.add_column("Score",     justify="center",  min_width=8)
    t.add_column("Grade",     justify="center",  min_width=5)
    t.add_column("Signal",    justify="center",  min_width=11)
    t.add_column("PE",        justify="right",   min_width=6)
    t.add_column("PS",        justify="right",   min_width=6)
    t.add_column("ROCE%",     justify="right",   min_width=7)
    t.add_column("OPM%",      justify="right",   min_width=6)
    t.add_column("NPM%",      justify="right",   min_width=6)
    t.add_column("FCF%",      justify="right",   min_width=6)
    t.add_column("Sales Gr%", justify="right",   min_width=8)
    t.add_column("CCC days",  justify="right",   min_width=8)
    t.add_column("ROE%",      justify="right",   min_width=6)
    t.add_column("EPS ₹",     justify="right",   min_width=7)
    t.add_column("Promo%",    justify="right",   min_width=7)

    for rank, s in enumerate(show, 1):
        cap_label = f"{s.exchange}/{s.cap[:3].upper()}"
        sig_style = _signal_style(s.signal)
        tot_color = _score_color(s.total, 100)
        ind       = s.indicators

        t.add_row(
            str(rank),
            s.symbol,
            s.name[:21],
            f"[dim]{cap_label}[/dim]",
            f"₹{s.price:,.0f}" if s.price else "—",
            f"[{tot_color}]{s.total:.0f}/100[/{tot_color}]",
            f"[{tot_color}]{s.grade}[/{tot_color}]",
            f"[{sig_style}]{s.signal}[/{sig_style}]",
            _fmt(ind.get("pe_ratio"),              ".1f"),
            _fmt(ind.get("ps_ratio"),              ".1f"),
            _fmt(ind.get("roce"),                  ".1f", "%"),
            _fmt(ind.get("operating_margin"),      ".1f", "%"),
            _fmt(ind.get("net_profit_margin"),     ".1f", "%"),
            _fmt(ind.get("fcf_margin"),            ".1f", "%"),
            _fmt(ind.get("sales_growth"),          ".1f", "%"),
            _fmt(ind.get("ccc"),                   ".0f", "d"),
            _fmt(ind.get("roe"),                   ".1f", "%"),
            _fmt(ind.get("net_eps"),               ".1f", ""),
            _fmt(ind.get("promoter_holding"),      ".1f", "%"),
        )

    console.print(t)

    console.print(
        "  Score guide: "
        "[bold green]80–100 = A+ Strong Buy[/bold green]  "
        "[green]70–79 = A Buy[/green]  "
        "[yellow]50–69 = B/C Watch[/yellow]  "
        "[red]< 50 = Avoid[/red]"
    )
    console.print()
    console.print(
        "  [dim]Promoter% = heldPercentInsiders proxy (approximate). "
        "Not available via free APIs: Change in Promoter Holding, Promoter Buying, "
        "Order Book, Segmental Revenue, Sales Breakup.[/dim]"
    )
    console.print()


# ─── CSV export ───────────────────────────────────────────────────────────────

def _rnd(v, n=2):
    try:
        import math
        return round(float(v), n) if not math.isnan(float(v)) else None
    except (TypeError, ValueError):
        return None


def save_csv(scores: list[ScoreBreakdown], output_dir: str):
    rows = []
    for s in sorted(scores, key=lambda x: x.total, reverse=True):
        ind = s.indicators
        rows.append({
            "Symbol":              s.symbol,
            "Name":                s.name,
            "Exchange":            s.exchange,
            "Cap":                 s.cap,
            "Price (₹)":           _rnd(s.price, 2),
            "Total Score":         _rnd(s.total, 1),
            "Grade":               s.grade,
            "Signal":              s.signal,
            # Score categories
            "Valuation /20":       _rnd(s.valuation, 1),
            "Profitability /25":   _rnd(s.profitability, 1),
            "Growth /20":          _rnd(s.growth, 1),
            "Efficiency /20":      _rnd(s.efficiency, 1),
            "Quality /15":         _rnd(s.quality, 1),
            # Raw fundamental ratios
            "P/E Ratio":                  _rnd(ind.get("pe_ratio"), 1),
            "Market Cap/Sales":           _rnd(ind.get("ps_ratio"), 2),
            "ROCE %":                     _rnd(ind.get("roce"), 1),
            "Operating Margin %":         _rnd(ind.get("operating_margin"), 1),
            "Net Profit Margin %":        _rnd(ind.get("net_profit_margin"), 1),
            "FCF Margin %":               _rnd(ind.get("fcf_margin"), 1),
            "Free Cash Flow (₹)":         _rnd(ind.get("fcf"), 0),
            "Net EPS (₹)":                _rnd(ind.get("net_eps"), 2),
            "Sales Growth % YoY":         _rnd(ind.get("sales_growth"), 1),
            "Capex/Sales %":              _rnd(ind.get("capex_sales"), 1),
            "Receivable/Sales %":         _rnd(ind.get("receivable_sales"), 1),
            "Receivable Days":            _rnd(ind.get("receivable_days"), 0),
            "Cash Conv. Cycle":           _rnd(ind.get("ccc"), 0),
            "ROE %":                      _rnd(ind.get("roe"), 1),
            "Promoter Holding % (proxy)": _rnd(ind.get("promoter_holding"), 1),
            # Not available via free APIs
            "Change in Promoter Holding": None,
            "Promoter Buying":            None,
            "Order Book":                 None,
            "Segmental Revenue":          None,
            "Sales Breakup":              None,
            "Error":                      s.error,
        })
    df = pd.DataFrame(rows)
    path = os.path.join(output_dir, "scores.csv")
    df.to_csv(path, index=False)
    console.print(f"  [green]CSV saved →[/green] {path}")
    return df


# ─── Excel export ─────────────────────────────────────────────────────────────

def save_excel(df: pd.DataFrame, output_dir: str):
    if not EXCEL:
        return
    path = os.path.join(output_dir, "scores.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Scores")
        ws = writer.sheets["Scores"]

        # Header style
        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Conditional colour for Total Score column (col F = index 6)
        score_col = 6
        for row in ws.iter_rows(min_row=2, min_col=score_col, max_col=score_col):
            for cell in row:
                try:
                    v = float(cell.value or 0)
                    if v >= 80:   color = "00B050"   # dark green
                    elif v >= 70: color = "92D050"   # light green
                    elif v >= 55: color = "FFEB9C"   # yellow
                    elif v >= 40: color = "FFCC00"   # amber
                    else:         color = "FF0000"   # red
                    cell.fill = PatternFill("solid", fgColor=color)
                except (TypeError, ValueError):
                    pass

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

    console.print(f"  [green]Excel saved →[/green] {path}")


# ─── Charts ───────────────────────────────────────────────────────────────────

def _signal_hex(signal: str) -> str:
    return {
        "STRONG BUY": "#27ae60",
        "BUY":        "#2ecc71",
        "WATCH":      "#f39c12",
        "NEUTRAL":    "#95a5a6",
        "AVOID":      "#e74c3c",
    }.get(signal, "#7f8c8d")


def save_top_chart(scores: list[ScoreBreakdown], output_dir: str, top_n: int = 20):
    if not MATPLOTLIB:
        return
    valid = sorted([s for s in scores if not s.error], key=lambda x: x.total, reverse=True)
    show  = valid[:top_n]

    labels  = [s.symbol for s in show]
    totals  = [s.total  for s in show]
    colors  = [_signal_hex(s.signal) for s in show]

    fig, ax = plt.subplots(figsize=(14, 7))
    bars = ax.barh(labels[::-1], totals[::-1], color=colors[::-1], edgecolor="white", linewidth=0.5)

    for bar, score in zip(bars, totals[::-1]):
        ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
                f"{score:.0f}", va="center", fontsize=9, fontweight="bold")

    ax.set_xlim(0, 110)
    ax.set_xlabel("Score (out of 100)", fontsize=11)
    ax.set_title(f"Top {top_n} Stocks — NSE/BSE Screener Score", fontsize=13, fontweight="bold")
    ax.axvline(75, color="#27ae60", linestyle="--", alpha=0.5, label="Strong Buy (75)")
    ax.axvline(60, color="#f39c12", linestyle="--", alpha=0.5, label="Buy (60)")

    legend = [
        mpatches.Patch(color="#27ae60", label="Strong Buy (≥75)"),
        mpatches.Patch(color="#2ecc71", label="Buy (60–74)"),
        mpatches.Patch(color="#f39c12", label="Watch (45–59)"),
        mpatches.Patch(color="#95a5a6", label="Neutral (30–44)"),
        mpatches.Patch(color="#e74c3c", label="Avoid (<30)"),
    ]
    ax.legend(handles=legend, loc="lower right", fontsize=9)
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()

    path = os.path.join(output_dir, "top20_chart.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    console.print(f"  [green]Top-20 chart →[/green] {path}")


def save_heatmap(scores: list[ScoreBreakdown], output_dir: str, top_n: int = 40):
    if not MATPLOTLIB:
        return
    valid = sorted([s for s in scores if not s.error], key=lambda x: x.total, reverse=True)
    show  = valid[:top_n]

    data = np.array([
        [s.valuation / 20, s.profitability / 25, s.growth / 20,
         s.efficiency / 20, s.quality / 15]
        for s in show
    ])
    labels_y = [s.symbol for s in show]
    labels_x = ["Valuation\n/20", "Profitability\n/25", "Growth\n/20",
                 "Efficiency\n/20", "Quality\n/15"]

    fig, ax = plt.subplots(figsize=(10, max(8, top_n * 0.35)))
    im = ax.imshow(data, cmap="RdYlGn", aspect="auto", vmin=0, vmax=1)

    ax.set_xticks(range(5))
    ax.set_xticklabels(labels_x, fontsize=10)
    ax.set_yticks(range(len(labels_y)))
    ax.set_yticklabels(labels_y, fontsize=8)

    for i in range(len(show)):
        for j, (val, max_v) in enumerate(zip(
            [show[i].valuation, show[i].profitability, show[i].growth,
             show[i].efficiency, show[i].quality],
            [20, 25, 20, 20, 15]
        )):
            ax.text(j, i, f"{val:.0f}", ha="center", va="center",
                    fontsize=7, color="black" if 0.3 < data[i,j] < 0.8 else "white")

    plt.colorbar(im, ax=ax, label="Score fraction (1.0 = max)")
    ax.set_title(f"Score Heatmap — Top {top_n} Stocks", fontsize=13, fontweight="bold")
    plt.tight_layout()

    path = os.path.join(output_dir, "heatmap.png")
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    console.print(f"  [green]Heatmap →[/green] {path}")


# ─── Watchlist export (used by simulation + live_signals) ────────────────────

# Shared file path — both modules import this constant to locate the file
WATCHLIST_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),   # Trading/
    "stock_screener", "results", "top_watchlist.json"
)

def save_top_watchlist(scores: list[ScoreBreakdown], top_n: int = 50):
    """
    Save the top N scored stocks to a shared JSON file that the simulation
    and live_signals modules read as their watchlist.
    """
    import json
    valid = [s for s in scores if not s.error and s.total > 0]
    valid.sort(key=lambda s: s.total, reverse=True)
    top = valid[:top_n]

    data = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "top_n":        top_n,
        "stocks": [
            {
                "symbol":   s.symbol,
                "name":     s.name,
                "exchange": s.exchange,
                "cap":      s.cap,
                "score":    round(s.total, 1),
                "grade":    s.grade,
                "signal":   s.signal,
            }
            for s in top
        ],
    }

    os.makedirs(os.path.dirname(WATCHLIST_PATH), exist_ok=True)
    with open(WATCHLIST_PATH, "w") as f:
        json.dump(data, f, indent=2)

    console.print(
        f"  [bold green]Watchlist saved →[/bold green] {WATCHLIST_PATH}  "
        f"[dim]({top_n} stocks — simulation & live_signals will use these)[/dim]"
    )


# ─── Master export function ───────────────────────────────────────────────────

def generate_report(scores: list[ScoreBreakdown], output_dir: str, top_n: int = 30):
    os.makedirs(output_dir, exist_ok=True)
    print_results(scores, top_n=top_n)
    df = save_csv(scores, output_dir)
    save_excel(df, output_dir)
    save_top_chart(scores, output_dir)
    save_heatmap(scores, output_dir)
    save_top_watchlist(scores, top_n=50)   # always save top 50 for other modules
    console.print()
